from contextlib import suppress
from datetime import UTC, datetime

from custom_data_toolkit.middleware.error_handler import (
    AppException,
    ConflictException,
    NotFoundException,
)
from custom_data_toolkit.models.customs_dict import (
    DICT_TEXT_MAX_LENGTH,
    CustomsDictMapping,
    CustomsDictSource,
    CustomsDictSyncStatus,
    normalize_dict_text,
    normalize_dict_type_code,
)
from custom_data_toolkit.repositories.customs_dict_repository import CustomsDictRepository
from custom_data_toolkit.repositories.customs_dict_type_repository import (
    CustomsDictTypeRepository,
)
from custom_data_toolkit.services.customs_dict_redis import (
    CUSTOMS_DICT_XLSX_HEADERS,
    IMPORT_MAX_ROWS,
    CustomsDictRedisStore,
    sanitize_redis_error,
)


class CustomsDictService:
    def __init__(
        self,
        repository: CustomsDictRepository,
        redis_store: CustomsDictRedisStore,
        type_repository: CustomsDictTypeRepository | None = None,
    ) -> None:
        self.repository = repository
        self.redis_store = redis_store
        self.type_repository = type_repository or CustomsDictTypeRepository(repository.session)

    def _require_enabled_type(self, dict_type: str) -> str:
        cleaned = normalize_dict_type_code(dict_type)
        if not cleaned:
            raise AppException("请选择字典类型。", error_code="CustomsDict.InvalidType")
        row = self.type_repository.get_by_code(cleaned)
        if row is None or not row.enabled:
            raise AppException(
                "字典类型不存在或已停用。",
                error_code="CustomsDict.InvalidType",
            )
        return row.code

    def _require_existing_type(self, dict_type: str) -> str:
        cleaned = normalize_dict_type_code(dict_type)
        if not cleaned:
            raise AppException("请选择字典类型。", error_code="CustomsDict.InvalidType")
        row = self.type_repository.get_by_code(cleaned)
        if row is None:
            raise AppException(
                "字典类型不存在。",
                error_code="CustomsDict.InvalidType",
            )
        return row.code

    def _type_label(self, code: str) -> str:
        row = self.type_repository.get_by_code(code)
        if row is not None:
            return row.name
        return code

    def list_page(
        self,
        *,
        dict_type: str | None,
        raw_value: str | None,
        standard_value: str | None,
        enabled: bool | None,
        page: int,
        page_size: int,
    ) -> tuple[list[CustomsDictMapping], int]:
        cleaned_type = None
        if dict_type is not None and dict_type.strip():
            # 列表筛选允许历史/停用类型 code
            cleaned_type = normalize_dict_type_code(dict_type)
        cleaned_raw = normalize_dict_text(raw_value) if raw_value else None
        cleaned_standard = normalize_dict_text(standard_value) if standard_value else None
        if cleaned_raw == "":
            cleaned_raw = None
        if cleaned_standard == "":
            cleaned_standard = None
        return self.repository.list_page(
            dict_type=cleaned_type,
            raw_value=cleaned_raw,
            standard_value=cleaned_standard,
            enabled=enabled,
            page=page,
            page_size=page_size,
        )

    def get(self, mapping_id: int) -> CustomsDictMapping:
        mapping = self.repository.get_by_id(mapping_id)
        if mapping is None:
            raise NotFoundException("未找到该字典映射。", error_code="CustomsDict.NotFound")
        return mapping

    def create(
        self,
        *,
        dict_type: str,
        raw_value: str,
        standard_value: str,
        actor_id: int | None,
        source: str = CustomsDictSource.MANUAL.value,
    ) -> CustomsDictMapping:
        cleaned_type = self._require_enabled_type(dict_type)
        cleaned_raw = self._require_text(raw_value, field="原始值")
        cleaned_standard = self._require_text(standard_value, field="标准值")
        existing = self.repository.get_by_type_raw(cleaned_type, cleaned_raw)
        if existing is not None:
            raise ConflictException(
                "该字典类型下原始值已存在。",
                error_code="CustomsDict.DuplicateRawValue",
            )
        now = datetime.now(UTC).replace(tzinfo=None)
        mapping = CustomsDictMapping(
            dict_type=cleaned_type,
            raw_value=cleaned_raw,
            standard_value=cleaned_standard,
            enabled=True,
            source=source,
            sync_status=CustomsDictSyncStatus.PENDING.value,
            created_by=actor_id,
            updated_by=actor_id,
            created_at=now,
            updated_at=now,
        )
        self.repository.add(mapping)
        self.repository.commit()
        self.repository.refresh(mapping)
        self._sync_mapping(mapping)
        return mapping

    def update_standard_value(
        self,
        mapping_id: int,
        *,
        standard_value: str,
        raw_value: str | None,
        actor_id: int | None,
    ) -> CustomsDictMapping:
        mapping = self.get(mapping_id)
        if raw_value is not None and normalize_dict_text(raw_value) != mapping.raw_value:
            raise AppException(
                "原始值创建后不可修改。",
                error_code="CustomsDict.RawValueImmutable",
            )
        mapping.standard_value = self._require_text(standard_value, field="标准值")
        mapping.updated_by = actor_id
        mapping.updated_at = datetime.now(UTC).replace(tzinfo=None)
        mapping.sync_status = CustomsDictSyncStatus.PENDING.value
        mapping.sync_error = None
        self.repository.commit()
        self.repository.refresh(mapping)
        self._sync_mapping(mapping)
        return mapping

    def set_enabled(
        self,
        mapping_id: int,
        *,
        enabled: bool,
        actor_id: int | None,
    ) -> CustomsDictMapping:
        mapping = self.get(mapping_id)
        mapping.enabled = enabled
        mapping.updated_by = actor_id
        mapping.updated_at = datetime.now(UTC).replace(tzinfo=None)
        mapping.sync_status = CustomsDictSyncStatus.PENDING.value
        mapping.sync_error = None
        self.repository.commit()
        self.repository.refresh(mapping)
        self._sync_mapping(mapping)
        return mapping

    def resync(self, mapping_id: int) -> CustomsDictMapping:
        mapping = self.get(mapping_id)
        self._sync_mapping(mapping)
        return mapping

    def batch_disable(
        self,
        mapping_ids: list[int],
        *,
        actor_id: int | None,
    ) -> dict[str, object]:
        mappings = self.repository.get_by_ids_for_update(mapping_ids)
        found_ids = {mapping.id for mapping in mappings}
        missing_ids = sorted(set(mapping_ids) - found_ids)
        if missing_ids:
            raise ConflictException(
                "部分映射已不存在，本次未停用任何记录，请刷新列表后重试。",
                error_code="BatchDelete.StaleSelection",
                details={"missingIds": missing_ids},
            )

        now = datetime.now(UTC).replace(tzinfo=None)
        for mapping in mappings:
            mapping.enabled = False
            mapping.updated_by = actor_id
            mapping.updated_at = now
            mapping.sync_status = CustomsDictSyncStatus.PENDING.value
            mapping.sync_error = None
        self.repository.commit()

        failed_ids: list[int] = []
        for mapping in mappings:
            self.repository.refresh(mapping)
            self._sync_mapping(mapping)
            self.repository.refresh(mapping)
            if mapping.sync_status != CustomsDictSyncStatus.SYNCED.value:
                failed_ids.append(mapping.id)

        return {
            "disabled": len(mappings),
            "sync_failed": len(failed_ids),
            "failed_ids": failed_ids,
        }

    def batch_resync(self, mapping_ids: list[int]) -> dict[str, object]:
        mappings = self.repository.get_by_ids_for_update(mapping_ids)
        found_ids = {mapping.id for mapping in mappings}
        missing_ids = sorted(set(mapping_ids) - found_ids)
        if missing_ids:
            raise ConflictException(
                "部分映射已不存在，本次未同步任何记录，请刷新列表后重试。",
                error_code="BatchDelete.StaleSelection",
                details={"missingIds": missing_ids},
            )

        synced = 0
        failed_ids: list[int] = []
        # 按请求顺序稳定返回 failedIds
        id_order = {mapping_id: index for index, mapping_id in enumerate(mapping_ids)}
        mappings_sorted = sorted(mappings, key=lambda item: id_order[item.id])
        for mapping in mappings_sorted:
            self._sync_mapping(mapping)
            self.repository.refresh(mapping)
            if mapping.sync_status == CustomsDictSyncStatus.SYNCED.value:
                synced += 1
            else:
                failed_ids.append(mapping.id)

        return {
            "synced": synced,
            "failed": len(failed_ids),
            "failed_ids": failed_ids,
            "total": len(mappings),
        }

    def list_missing(
        self,
        *,
        dict_type: str,
        raw_value: str | None,
        page: int,
        page_size: int,
    ) -> tuple[list[dict[str, object]], int]:
        cleaned_type = self._require_existing_type(dict_type)
        cleaned_raw = normalize_dict_text(raw_value) if raw_value else None
        if cleaned_raw == "":
            cleaned_raw = None
        try:
            items, total = self.redis_store.list_missing_page(
                dict_type=cleaned_type,
                raw_value=cleaned_raw,
                page=page,
                page_size=page_size,
            )
        except Exception as exc:  # noqa: BLE001
            raise AppException(
                f"读取缺失字典失败：{sanitize_redis_error(exc)}",
                error_code="CustomsDict.MissingReadFailed",
            ) from exc
        label = self._type_label(cleaned_type)
        return [
            {
                "dict_type": cleaned_type,
                "dict_type_label": label,
                "raw_value": member,
                "occurrence_count": int(score),
            }
            for member, score in items
        ], total

    def handle_missing(
        self,
        *,
        dict_type: str,
        raw_value: str,
        standard_value: str,
        actor_id: int | None,
    ) -> CustomsDictMapping:
        cleaned_type = self._require_enabled_type(dict_type)
        cleaned_raw = self._require_text(raw_value, field="原始值")
        cleaned_standard = self._require_text(standard_value, field="标准值")
        existing = self.repository.get_by_type_raw(cleaned_type, cleaned_raw)
        if existing is not None:
            raise ConflictException(
                "该字典类型下原始值已存在。",
                error_code="CustomsDict.DuplicateRawValue",
            )
        now = datetime.now(UTC).replace(tzinfo=None)
        mapping = CustomsDictMapping(
            dict_type=cleaned_type,
            raw_value=cleaned_raw,
            standard_value=cleaned_standard,
            enabled=True,
            source=CustomsDictSource.MISSING.value,
            sync_status=CustomsDictSyncStatus.PENDING.value,
            created_by=actor_id,
            updated_by=actor_id,
            created_at=now,
            updated_at=now,
        )
        self.repository.add(mapping)
        self.repository.commit()
        self.repository.refresh(mapping)
        self._sync_mapping(mapping)
        return mapping

    def export_missing_xlsx(
        self,
        *,
        dict_type: str,
        raw_value: str | None,
    ) -> bytes:
        from io import BytesIO

        from openpyxl import Workbook

        cleaned_type = self._require_existing_type(dict_type)
        cleaned_raw = normalize_dict_text(raw_value) if raw_value else None
        if cleaned_raw == "":
            cleaned_raw = None
        try:
            items = self.redis_store.list_missing_all(
                dict_type=cleaned_type,
                raw_value=cleaned_raw,
            )
        except Exception as exc:  # noqa: BLE001
            raise AppException(
                f"导出缺失字典失败：{sanitize_redis_error(exc)}",
                error_code="CustomsDict.MissingReadFailed",
            ) from exc

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "missing"
        sheet.append(list(CUSTOMS_DICT_XLSX_HEADERS))
        label = self._type_label(cleaned_type)
        for member, score in items:
            sheet.append([cleaned_type, label, member, int(score), "", ""])
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def export_mappings_xlsx(
        self,
        *,
        dict_type: str | None,
        raw_value: str | None,
        standard_value: str | None,
        enabled: bool | None,
    ) -> bytes:
        from io import BytesIO

        from openpyxl import Workbook

        cleaned_type = None
        if dict_type is not None and dict_type.strip():
            cleaned_type = normalize_dict_type_code(dict_type)
        cleaned_raw = normalize_dict_text(raw_value) if raw_value else None
        cleaned_standard = normalize_dict_text(standard_value) if standard_value else None
        if cleaned_raw == "":
            cleaned_raw = None
        if cleaned_standard == "":
            cleaned_standard = None

        mappings = self.repository.list_all_filtered(
            dict_type=cleaned_type,
            raw_value=cleaned_raw,
            standard_value=cleaned_standard,
            enabled=enabled,
        )
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "mappings"
        sheet.append(list(CUSTOMS_DICT_XLSX_HEADERS))
        for mapping in mappings:
            sheet.append(
                [
                    mapping.dict_type,
                    self._type_label(mapping.dict_type),
                    mapping.raw_value,
                    "",
                    mapping.standard_value,
                    "",
                ]
            )
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def import_template_xlsx(self) -> bytes:
        from io import BytesIO

        from openpyxl import Workbook

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "mappings"
        sheet.append(list(CUSTOMS_DICT_XLSX_HEADERS))
        buffer = BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def import_mappings_xlsx(
        self,
        *,
        content: bytes,
        actor_id: int | None,
    ) -> dict[str, object]:
        from io import BytesIO

        from openpyxl import load_workbook

        try:
            workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001
            raise AppException(
                "无法解析 Excel 文件，请上传有效的 xlsx。",
                error_code="CustomsDict.ImportInvalidFile",
            ) from exc

        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration as exc:
            raise AppException(
                "导入文件为空。",
                error_code="CustomsDict.ImportEmpty",
            ) from exc

        header_map = self._xlsx_header_map(header_row)
        created = 0
        updated = 0
        failed = 0
        errors: list[dict[str, object]] = []
        data_rows = 0

        for excel_row_number, row in enumerate(rows, start=2):
            if self._xlsx_row_empty(row):
                continue
            data_rows += 1
            if data_rows > IMPORT_MAX_ROWS:
                errors.append(
                    {
                        "row": excel_row_number,
                        "message": f"超过单次导入上限 {IMPORT_MAX_ROWS} 行。",
                    }
                )
                failed += 1
                break
            try:
                dict_type = self._xlsx_cell(row, header_map["字典类型编码"])
                raw_value = self._xlsx_cell(row, header_map["原始值"])
                standard_value = self._xlsx_cell(row, header_map["标准值"])
                cleaned_type = self._require_enabled_type(dict_type)
                cleaned_raw = self._require_text(raw_value, field="原始值")
                cleaned_standard = self._require_text(standard_value, field="标准值")
                existing = self.repository.get_by_type_raw(cleaned_type, cleaned_raw)
                if existing is None:
                    self.create(
                        dict_type=cleaned_type,
                        raw_value=cleaned_raw,
                        standard_value=cleaned_standard,
                        actor_id=actor_id,
                        source=CustomsDictSource.IMPORT.value,
                    )
                    created += 1
                else:
                    assert existing.id is not None
                    self.update_standard_value(
                        existing.id,
                        standard_value=cleaned_standard,
                        raw_value=None,
                        actor_id=actor_id,
                    )
                    updated += 1
            except AppException as exc:
                failed += 1
                errors.append({"row": excel_row_number, "message": exc.message})
            except Exception as exc:  # noqa: BLE001
                failed += 1
                errors.append({"row": excel_row_number, "message": str(exc)[:200]})

        return {
            "created": created,
            "updated": updated,
            "failed": failed,
            "errors": errors,
        }

    @staticmethod
    def _xlsx_header_map(header_row: tuple[object, ...] | list[object]) -> dict[str, int]:
        normalized = {
            str(cell).strip(): index
            for index, cell in enumerate(header_row)
            if cell is not None and str(cell).strip()
        }
        missing = [name for name in ("字典类型编码", "原始值", "标准值") if name not in normalized]
        if missing:
            raise AppException(
                f"表头缺少列：{', '.join(missing)}。请使用与导出相同的模板。",
                error_code="CustomsDict.ImportBadHeader",
            )
        return normalized

    @staticmethod
    def _xlsx_cell(row: tuple[object, ...] | list[object], index: int) -> str:
        if index >= len(row):
            return ""
        value = row[index]
        if value is None:
            return ""
        return str(value)

    @staticmethod
    def _xlsx_row_empty(row: tuple[object, ...] | list[object] | None) -> bool:
        if row is None:
            return True
        return all(cell is None or str(cell).strip() == "" for cell in row)

    def replay_sync(self, *, dict_type: str) -> dict[str, int]:
        cleaned_type = self._require_existing_type(dict_type)
        mappings = self.repository.list_by_dict_type(cleaned_type)
        synced = 0
        failed = 0
        for mapping in mappings:
            self._sync_mapping(mapping)
            self.repository.refresh(mapping)
            if mapping.sync_status == CustomsDictSyncStatus.SYNCED.value:
                synced += 1
            else:
                failed += 1
        return {"synced": synced, "failed": failed, "total": len(mappings)}

    def _sync_mapping(self, mapping: CustomsDictMapping) -> None:
        try:
            if mapping.enabled:
                self.redis_store.put(
                    dict_type=mapping.dict_type,
                    raw_value=mapping.raw_value,
                    standard_value=mapping.standard_value,
                )
            else:
                self.redis_store.remove(
                    dict_type=mapping.dict_type,
                    raw_value=mapping.raw_value,
                )
        except Exception as exc:  # noqa: BLE001 — Redis 故障不回滚 MySQL
            mapping.sync_status = CustomsDictSyncStatus.FAILED.value
            mapping.sync_error = sanitize_redis_error(exc)
            self.repository.commit()
            self.repository.refresh(mapping)
            return
        mapping.sync_status = CustomsDictSyncStatus.SYNCED.value
        mapping.sync_error = None
        mapping.last_synced_at = datetime.now(UTC).replace(tzinfo=None)
        self.repository.commit()
        self.repository.refresh(mapping)
        if mapping.enabled:
            with suppress(Exception):
                self.redis_store.remove_missing(
                    dict_type=mapping.dict_type,
                    raw_value=mapping.raw_value,
                )

    @staticmethod
    def _require_text(value: str, *, field: str) -> str:
        cleaned = normalize_dict_text(value)
        if not cleaned:
            raise AppException(f"请填写{field}。", error_code="CustomsDict.EmptyValue")
        if len(cleaned) > DICT_TEXT_MAX_LENGTH:
            raise AppException(
                f"{field}长度不能超过 {DICT_TEXT_MAX_LENGTH}。",
                error_code="CustomsDict.ValueTooLong",
            )
        return cleaned
