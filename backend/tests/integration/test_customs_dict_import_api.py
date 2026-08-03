"""标准字典 xlsx 导出/导入集成测试。"""

from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO

import fakeredis
import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import text
from sqlmodel import Session

from custom_data_toolkit.config.settings import settings
from custom_data_toolkit.db.engine import engine
from custom_data_toolkit.deps import SessionDep
from custom_data_toolkit.main import app
from custom_data_toolkit.repositories.customs_dict_repository import CustomsDictRepository
from custom_data_toolkit.routers.auth import CSRF_COOKIE_NAME
from custom_data_toolkit.routers.customs_dict import get_customs_dict_service
from custom_data_toolkit.services.customs_dict_redis import (
    CUSTOMS_DICT_XLSX_HEADERS,
    CustomsDictRedisStore,
    formal_dict_key,
)
from custom_data_toolkit.services.customs_dict_service import CustomsDictService


def _db_ready() -> bool:
    try:
        with Session(engine) as session:
            session.exec(text("SELECT 1 FROM admin_users LIMIT 1"))
            session.exec(text("SELECT 1 FROM customs_dict_mapping LIMIT 1"))
        return True
    except Exception:
        return False


pytestmark = pytest.mark.skipif(
    not _db_ready(),
    reason="MySQL admin/customs_dict_mapping tables not ready",
)


@pytest.fixture()
def fake_redis() -> fakeredis.FakeRedis:
    client = fakeredis.FakeRedis(decode_responses=True)
    yield client
    client.flushall()


@pytest.fixture()
def client(fake_redis: fakeredis.FakeRedis):
    def _override(session: SessionDep) -> CustomsDictService:
        return CustomsDictService(
            CustomsDictRepository(session),
            CustomsDictRedisStore(fake_redis),
        )

    app.dependency_overrides[get_customs_dict_service] = _override
    with TestClient(app) as test_client:
        yield test_client
    app.dependency_overrides.pop(get_customs_dict_service, None)


def _login(client: TestClient) -> dict[str, str]:
    csrf = client.get("/api/v1/auth/csrf")
    assert csrf.status_code == 200
    headers = {"X-CSRF-Token": csrf.json()["csrfToken"]}
    login = client.post(
        "/api/v1/auth/login",
        json={
            "username": settings.admin_bootstrap_username,
            "password": settings.admin_bootstrap_password,
        },
        headers=headers,
    )
    assert login.status_code == 200
    token = client.cookies.get(CSRF_COOKIE_NAME)
    assert token
    return {"X-CSRF-Token": token}


def _xlsx_bytes(rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    for row in rows:
        sheet.append(row)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_export_import_round_trip_updates_standard_value(
    client: TestClient,
    fake_redis: fakeredis.FakeRedis,
) -> None:
    headers = _login(client)
    suffix = datetime.now(UTC).strftime("%H%M%S%f")
    raw = f"imp-rt-{suffix}"
    created = client.post(
        "/api/v1/customs-dict/mappings",
        json={"dictType": "country", "rawValue": raw, "standardValue": "OLD"},
        headers=headers,
    )
    assert created.status_code == 201, created.text

    exported = client.get(
        "/api/v1/customs-dict/mappings/export",
        params={"dictType": "country", "rawValue": raw},
        headers=headers,
    )
    assert exported.status_code == 200
    workbook = load_workbook(BytesIO(exported.content))
    rows = list(workbook.active.iter_rows(values_only=True))
    assert rows[0] == CUSTOMS_DICT_XLSX_HEADERS
    data_row = next(row for row in rows[1:] if row[2] == raw)
    assert data_row[4] == "OLD"

    # 改标准值再导入
    payload = _xlsx_bytes(
        [
            list(CUSTOMS_DICT_XLSX_HEADERS),
            ["country", "国家", raw, "", "NEW", ""],
        ]
    )
    imported = client.post(
        "/api/v1/customs-dict/mappings/import",
        headers=headers,
        files={
            "file": (
                "import.xlsx",
                payload,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
    )
    assert imported.status_code == 200, imported.text
    body = imported.json()
    assert body["created"] == 0
    assert body["updated"] == 1
    assert body["failed"] == 0
    assert fake_redis.hget(formal_dict_key("country"), raw) == "NEW"


def test_import_creates_and_reports_bad_rows(
    client: TestClient,
    fake_redis: fakeredis.FakeRedis,
) -> None:
    headers = _login(client)
    suffix = datetime.now(UTC).strftime("%H%M%S%f")
    raw_ok = f"imp-ok-{suffix}"
    payload = _xlsx_bytes(
        [
            list(CUSTOMS_DICT_XLSX_HEADERS),
            ["country", "国家", raw_ok, 3, "STD", "备注"],
            ["badtype", "x", "x", "", "y", ""],
            ["country", "国家", "", "", "STD2", ""],
        ]
    )
    imported = client.post(
        "/api/v1/customs-dict/mappings/import",
        headers=headers,
        files={
            "file": (
                "import.xlsx",
                payload,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ),
        },
    )
    assert imported.status_code == 200, imported.text
    body = imported.json()
    assert body["created"] == 1
    assert body["updated"] == 0
    assert body["failed"] == 2
    assert len(body["errors"]) == 2

    listed = client.get(
        "/api/v1/customs-dict/mappings",
        params={"dictType": "country", "rawValue": raw_ok},
        headers=headers,
    )
    assert listed.status_code == 200
    item = next(i for i in listed.json()["items"] if i["rawValue"] == raw_ok)
    assert item["source"] == "import"
    assert item["standardValue"] == "STD"
    assert fake_redis.hget(formal_dict_key("country"), raw_ok) == "STD"


def test_import_template_headers(client: TestClient) -> None:
    headers = _login(client)
    response = client.get("/api/v1/customs-dict/mappings/import-template", headers=headers)
    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    assert list(workbook.active.iter_rows(values_only=True))[0] == CUSTOMS_DICT_XLSX_HEADERS
