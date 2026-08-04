"""海关缺失字典 API 集成测试（fakeredis）。"""

from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO

import fakeredis
import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import text
from sqlmodel import Session

from custom_data_toolkit.config.settings import settings
from custom_data_toolkit.db.engine import engine
from custom_data_toolkit.deps import SessionDep
from custom_data_toolkit.main import app
from custom_data_toolkit.repositories.customs_dict_repository import CustomsDictRepository
from custom_data_toolkit.routers.auth import CSRF_COOKIE_NAME
from custom_data_toolkit.routers.customs_dict import get_customs_dict_service
from custom_data_toolkit.services.customs_dict_redis import (
    CustomsDictRedisStore,
    formal_dict_key,
    missing_dict_key,
)
from custom_data_toolkit.services.customs_dict_service import CustomsDictService


def _db_ready() -> bool:
    try:
        with Session(engine) as session:
            session.exec(text("SELECT 1 FROM admin_users LIMIT 1"))
            session.exec(text("SELECT 1 FROM customs_dict_mapping LIMIT 1"))
        return True
    except Exception:
        return False


pytestmark = pytest.mark.skipif(
    not _db_ready(),
    reason="MySQL admin/customs_dict_mapping tables not ready",
)


@pytest.fixture()
def fake_redis() -> fakeredis.FakeRedis:
    client = fakeredis.FakeRedis(decode_responses=True)
    yield client
    client.flushall()


@pytest.fixture()
def client(fake_redis: fakeredis.FakeRedis):
    def _override(session: SessionDep) -> CustomsDictService:
        return CustomsDictService(
            CustomsDictRepository(session),
            CustomsDictRedisStore(fake_redis),
        )

    app.dependency_overrides[get_customs_dict_service] = _override
    with TestClient(app) as test_client:
        yield test_client
    app.dependency_overrides.pop(get_customs_dict_service, None)


def _login(client: TestClient) -> dict[str, str]:
    csrf = client.get("/api/v1/auth/csrf")
    assert csrf.status_code == 200
    headers = {"X-CSRF-Token": csrf.json()["csrfToken"]}
    login = client.post(
        "/api/v1/auth/login",
        json={
            "username": settings.admin_bootstrap_username,
            "password": settings.admin_bootstrap_password,
        },
        headers=headers,
    )
    assert login.status_code == 200
    token = client.cookies.get(CSRF_COOKIE_NAME)
    assert token
    return {"X-CSRF-Token": token}


def test_missing_list_handle_export_and_zrem(
    client: TestClient,
    fake_redis: fakeredis.FakeRedis,
) -> None:
    headers = _login(client)
    suffix = datetime.now(UTC).strftime("%H%M%S%f")
    raw_hi = f"高发缺失-{suffix}"
    raw_lo = f"低发缺失-{suffix}"
    key = missing_dict_key("country")
    fake_redis.zadd(key, {raw_lo: 2, raw_hi: 9})

    listed = client.get(
        "/api/v1/customs-dict/missing",
        params={"dictType": "country"},
        headers=headers,
    )
    assert listed.status_code == 200, listed.text
    body = listed.json()
    assert body["total"] >= 2
    assert body["items"][0]["rawValue"] == raw_hi
    assert body["items"][0]["occurrenceCount"] == 9

    handled = client.post(
        "/api/v1/customs-dict/missing/handle",
        json={"dictType": "country", "rawValue": raw_hi, "standardValue": "CHN"},
        headers=headers,
    )
    assert handled.status_code == 200, handled.text
    assert handled.json()["source"] == "missing"
    assert handled.json()["syncStatus"] == "synced"
    assert fake_redis.hget(formal_dict_key("country"), raw_hi) == "CHN"
    assert fake_redis.zscore(key, raw_hi) is None
    assert fake_redis.zscore(key, raw_lo) == 2

    dup = client.post(
        "/api/v1/customs-dict/missing/handle",
        json={"dictType": "country", "rawValue": raw_hi, "standardValue": "CHN"},
        headers=headers,
    )
    assert dup.status_code == 409

    exported = client.get(
        "/api/v1/customs-dict/missing/export",
        params={"dictType": "country"},
        headers=headers,
    )
    assert exported.status_code == 200
    assert fake_redis.zscore(key, raw_lo) == 2
    workbook = load_workbook(BytesIO(exported.content))
    rows = list(workbook.active.iter_rows(values_only=True))
    assert rows[0][:4] == ("字典类型编码", "字典类型名称", "原始值", "出现次数")
    assert any(row[2] == raw_lo for row in rows[1:])


def test_missing_list_all_types_when_dict_type_omitted(
    client: TestClient,
    fake_redis: fakeredis.FakeRedis,
) -> None:
    headers = _login(client)
    suffix = datetime.now(UTC).strftime("%H%M%S%f")
    country_raw = f"全国缺失-{suffix}"
    continent_raw = f"全洲缺失-{suffix}"
    fake_redis.zadd(missing_dict_key("country"), {country_raw: 7})
    fake_redis.zadd(missing_dict_key("continent"), {continent_raw: 11})

    listed = client.get("/api/v1/customs-dict/missing", headers=headers)
    assert listed.status_code == 200, listed.text
    body = listed.json()
    raw_values = {item["rawValue"] for item in body["items"]}
    assert country_raw in raw_values
    assert continent_raw in raw_values
    assert body["items"][0]["rawValue"] == continent_raw
    assert body["items"][0]["occurrenceCount"] == 11

    exported = client.get("/api/v1/customs-dict/missing/export", headers=headers)
    assert exported.status_code == 200
    workbook = load_workbook(BytesIO(exported.content))
    rows = list(workbook.active.iter_rows(values_only=True))
    assert any(row[2] == country_raw for row in rows[1:])
    assert any(row[2] == continent_raw for row in rows[1:])


def test_missing_suggestions_prefix(
    client: TestClient,
    fake_redis: fakeredis.FakeRedis,
) -> None:
    headers = _login(client)
    suffix = datetime.now(UTC).strftime("%H%M%S%f")
    raw = f"SuggestMiss-{suffix}"
    key = missing_dict_key("country")
    fake_redis.zadd(key, {raw: 3, f"Other-{suffix}": 1})

    suggestions = client.get(
        "/api/v1/customs-dict/missing/suggestions",
        params={"dictType": "country", "prefix": "SuggestMiss"},
        headers=headers,
    )
    assert suggestions.status_code == 200
    body = suggestions.json()
    assert len(body) <= 10
    assert any(item["rawValue"] == raw and item["occurrenceCount"] == 3 for item in body)

    cross = client.get(
        "/api/v1/customs-dict/missing/suggestions",
        params={"prefix": "SuggestMiss"},
        headers=headers,
    )
    assert cross.status_code == 200
    assert any(item["rawValue"] == raw for item in cross.json())
